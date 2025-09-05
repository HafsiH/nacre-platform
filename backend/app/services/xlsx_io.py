from typing import Iterator, Tuple, List, Dict, Any
from openpyxl import load_workbook


def _load_wb(path: str):
    return load_workbook(path, read_only=True, data_only=True)


def preview_xlsx(path: str, limit: int = 20) -> Tuple[List[str], List[Dict[str, Any]]]:
    wb = _load_wb(path)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return [], []
    cols = [str(h) if h is not None else "" for h in header]
    rows: List[Dict[str, Any]] = []
    for i, r in enumerate(rows_iter):
        if i >= limit:
            break
        rec = {cols[j]: (r[j] if j < len(r) else None) for j in range(len(cols))}
        rows.append({k: _to_str(v) for k, v in rec.items()})
    return cols, rows


def iterate_xlsx(path: str) -> Iterator[Dict[str, Any]]:
    wb = _load_wb(path)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return
    cols = [str(h) if h is not None else "" for h in header]
    for r in rows_iter:
        rec = {cols[j]: (r[j] if j < len(r) else None) for j in range(len(cols))}
        yield {k: _to_str(v) for k, v in rec.items()}


def count_xlsx_rows(path: str) -> int:
    wb = _load_wb(path)
    ws = wb.active
    # Data rows excluding header
    count = max(0, ws.max_row - 1)
    return count


def _to_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v)

